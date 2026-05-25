from io import BytesIO
import hashlib
import json
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

from app.repositories.auth_code_repository import AuthCodeRepository


class ImportValidationError(ValueError):
    def __init__(self, message: str, errors: list[str]) -> None:
        super().__init__(message)
        self.errors = errors


REQUIRED_PAYLOAD_FIELDS = ("did", "license")
HEADER_PID_CANDIDATES = {"pid", "product_pid", "产品pid", "产品id"}
HEADER_BATCH_CANDIDATES = {"batch", "source_batch", "批次"}
DISPLAY_PRIORITY = ("did", "auth_code", "license", "code")


class ExcelService:
    def __init__(self, repository: AuthCodeRepository) -> None:
        self.repository = repository

    def import_codes(self, file_stream: BinaryIO, default_pid: str | None = None) -> dict:
        rows = self._parse_rows(file_stream, default_pid)
        if not rows:
            raise ValueError("Excel 中没有可导入的授权码数据")
        result = self.repository.bulk_insert_codes(rows)
        result["total_rows"] = len(rows)
        return result

    def build_allocations_workbook(self) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "assigned_codes"
        rows = self.repository.list_assigned_codes_for_export()
        payload_keys = self._collect_payload_keys(rows)
        sheet.append(["PID", "DID", "MAC", "批次", "分配时间", "导入时间", *payload_keys])

        for item in rows:
            sheet.append(
                [
                    item["pid"],
                    item["did"],
                    item["assigned_mac"],
                    item["source_batch"],
                    item["assigned_at"],
                    item["created_at"],
                    *[item["payload"].get(key, "") for key in payload_keys],
                ]
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def build_logs_workbook(self, action: str = "all", search: str = "") -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "request_logs"
        rows = self.repository.list_logs_for_export(action=action, search=search)
        payload_keys = self._collect_payload_keys(rows)
        sheet.append(
            ["时间", "PID", "MAC", "动作", "DID", "说明", "来源IP", "载荷JSON", *payload_keys]
        )

        for item in rows:
            sheet.append(
                [
                    item["created_at"],
                    item["pid"],
                    item["mac"],
                    item["action"],
                    item["code"],
                    item["message"],
                    item["client_ip"],
                    item.get("payload_json") or "",
                    *[(item.get("payload") or {}).get(key, "") for key in payload_keys],
                ]
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def build_import_template(self) -> BytesIO:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "auth_codes_template"
        sheet.append(["pid", "did", "license", "source_batch"])
        sheet.append(["P1001", "DID-DEMO-001", "LICENSE-DEMO-001", "BATCH-01"])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _parse_rows(
        self,
        file_stream: BinaryIO,
        default_pid: str | None,
    ) -> list[dict[str, str]]:
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []

        first_row = list(values[0])
        header_map = self._match_headers(first_row)
        has_header = self._looks_like_header_row(first_row, header_map)

        data_rows = values[1:] if has_header else values
        parsed_rows: list[dict[str, str]] = []
        errors: list[str] = []
        start_row_number = 2 if has_header else 1
        for offset, row in enumerate(data_rows):
            current = list(row)
            row_number = start_row_number + offset
            try:
                if has_header:
                    parsed = self._parse_structured_row(current, first_row, default_pid)
                else:
                    parsed = self._parse_simple_row(current, default_pid)
            except ValueError as exc:
                errors.append(f"第 {row_number} 行：{exc}")
                continue
            if parsed:
                parsed_rows.append(parsed)
        if errors:
            raise ImportValidationError("导入校验失败，请先修正以下问题后再重试", errors)
        return parsed_rows

    @staticmethod
    def _match_headers(first_row: list[object]) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, value in enumerate(first_row):
            normalized = str(value or "").strip().lower()
            if normalized in HEADER_PID_CANDIDATES:
                header_map["pid"] = index
            if normalized in HEADER_BATCH_CANDIDATES:
                header_map["batch"] = index
            if normalized in REQUIRED_PAYLOAD_FIELDS:
                header_map[normalized] = index
        return header_map

    @staticmethod
    def _extract_cell(row: list[object], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    @staticmethod
    def _looks_like_header_row(first_row: list[object], header_map: dict[str, int]) -> bool:
        if header_map:
            return True
        tokens = [str(cell or "").strip() for cell in first_row if str(cell or "").strip()]
        if not tokens:
            return False
        alpha_like = 0
        for token in tokens:
            simplified = token.replace("_", "").replace(" ", "")
            if simplified.isalpha():
                alpha_like += 1
        return alpha_like >= max(1, len(tokens) - 1)

    def _parse_structured_row(
        self,
        row: list[object],
        headers: list[object],
        default_pid: str | None,
    ) -> dict[str, str] | None:
        payload: dict[str, str] = {}
        pid = (default_pid or "").strip()
        batch = ""

        for index, raw_header in enumerate(headers):
            header = str(raw_header or "").strip()
            if not header:
                continue
            value = self._extract_cell(row, index)
            normalized = header.lower()
            if normalized in HEADER_PID_CANDIDATES:
                pid = value or pid
                continue
            if normalized in HEADER_BATCH_CANDIDATES:
                batch = value
                continue
            if value:
                payload[header] = value

        if not payload:
            return None
        if not pid:
            raise ValueError("导入数据缺少 PID，请在 Excel 中提供 pid 列或填写默认 PID")

        missing_fields = [
            field for field in REQUIRED_PAYLOAD_FIELDS if not self._get_payload_value(payload, field)
        ]
        if missing_fields:
            raise ValueError(
                f"导入数据缺少必填字段：{', '.join(missing_fields)}。每条记录必须同时包含 did 和 license"
            )

        did = self._get_payload_value(payload, "did")
        license_value = self._get_payload_value(payload, "license")
        display_code = self._pick_display_code(payload)
        payload_json = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        return {
            "pid": pid,
            "did": did,
            "license": license_value,
            "code": display_code,
            "payload_json": payload_json,
            "payload_hash": hashlib.sha256(payload_json.encode("utf-8")).hexdigest(),
            "source_batch": batch or None,
        }

    def _parse_simple_row(
        self,
        row: list[object],
        default_pid: str | None,
    ) -> dict[str, str] | None:
        raise ValueError("导入文件必须包含表头，并至少提供 did 与 license 两列")

    @staticmethod
    def _pick_display_code(payload: dict[str, str]) -> str:
        lower_map = {key.lower(): value for key, value in payload.items()}
        for key in DISPLAY_PRIORITY:
            if lower_map.get(key):
                return lower_map[key]
        return next(iter(payload.values()))

    @staticmethod
    def _get_payload_value(payload: dict[str, str], field_name: str) -> str:
        if field_name in payload and payload[field_name]:
            return payload[field_name]
        lowered = field_name.lower()
        for key, value in payload.items():
            if key.lower() == lowered and value:
                return value
        return ""

    @staticmethod
    def _collect_payload_keys(rows: list[dict]) -> list[str]:
        seen: set[str] = set()
        keys: list[str] = []
        for preferred in DISPLAY_PRIORITY:
            for item in rows:
                for key in item.get("payload") or {}:
                    if key.lower() == preferred and key not in seen:
                        seen.add(key)
                        keys.append(key)
        for item in rows:
            for key in item.get("payload") or {}:
                if key not in seen:
                    seen.add(key)
                    keys.append(key)
        return keys
