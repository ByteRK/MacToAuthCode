import os
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from app import create_app
from app.config import Settings, load_settings


class PlatformTestCase(TestCase):
    def create_test_app(self, temp_dir: str):
        settings = Settings(
            app_name="Test",
            host="127.0.0.1",
            port=18080,
            admin_username="admin",
            admin_password="password",
            secret_key="test-secret",
            data_dir=Path(temp_dir),
        )
        return create_app(settings=settings)

    def test_load_settings_supports_request_ip_whitelist_config(self):
        with TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "config.json"
            config_path.write_text(
                """
                {
                  "app_name": "Test",
                  "host": "0.0.0.0",
                  "port": 8080,
                  "admin_username": "admin",
                  "admin_password": "password",
                  "secret_key": "secret",
                  "data_dir": "data",
                  "request_ip_whitelist": {
                    "enabled": true,
                    "allowed_ips": ["192.168.1.10", "192.168.1.0/24"]
                  }
                }
                """,
                encoding="utf-8",
            )
            original_config_file = os.environ.get("AUTH_PLATFORM_CONFIG_FILE")
            try:
                os.environ["AUTH_PLATFORM_CONFIG_FILE"] = str(config_path)
                settings = load_settings()
            finally:
                if original_config_file is None:
                    os.environ.pop("AUTH_PLATFORM_CONFIG_FILE", None)
                else:
                    os.environ["AUTH_PLATFORM_CONFIG_FILE"] = original_config_file

            self.assertTrue(settings.request_ip_whitelist_enabled)
            self.assertEqual(
                settings.request_ip_whitelist,
                ("192.168.1.10", "192.168.1.0/24"),
            )

    def test_distribution_is_scoped_by_pid(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-P1-001",
                        "license": "P1-001",
                        "code": "DID-P1-001",
                        "payload_json": "{\"did\": \"DID-P1-001\", \"license\": \"P1-001\"}",
                        "payload_hash": "hash-p1",
                        "source_batch": "B1",
                    },
                    {
                        "pid": "P2",
                        "did": "DID-P2-001",
                        "license": "P2-001",
                        "code": "DID-P2-001",
                        "payload_json": "{\"did\": \"DID-P2-001\", \"license\": \"P2-001\"}",
                        "payload_hash": "hash-p2",
                        "source_batch": "B2",
                    },
                ]
            )
            client = app.test_client()

            first = client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            second = client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            third = client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P2"},
            )

            self.assertEqual(first.status_code, 200)
            self.assertEqual(first.get_json()["data"]["display_code"], "DID-P1-001")
            self.assertEqual(first.get_json()["data"]["payload"]["license"], "P1-001")
            self.assertEqual(second.get_json()["data"]["mode"], "reused")
            self.assertEqual(third.get_json()["data"]["display_code"], "DID-P2-001")

    def test_device_authorize_rejects_requests_outside_ip_whitelist(self):
        with TemporaryDirectory() as temp_dir:
            settings = Settings(
                app_name="Test",
                host="127.0.0.1",
                port=18080,
                admin_username="admin",
                admin_password="password",
                secret_key="test-secret",
                data_dir=Path(temp_dir),
                request_ip_whitelist_enabled=True,
                request_ip_whitelist=("192.168.1.0/24",),
            )
            app = create_app(settings=settings)
            repo = app.extensions["auth_code_repository"]
            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-P1-001",
                        "license": "P1-001",
                        "code": "DID-P1-001",
                        "payload_json": "{\"did\": \"DID-P1-001\", \"license\": \"P1-001\"}",
                        "payload_hash": "hash-p1",
                        "source_batch": "B1",
                    },
                ]
            )
            client = app.test_client()

            response = client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
                environ_overrides={"REMOTE_ADDR": "10.0.0.2"},
            )
            payload = response.get_json()

            self.assertEqual(response.status_code, 403)
            self.assertFalse(payload["success"])
            self.assertIn("IP 不在白名单", payload["message"])

            codes = repo.list_codes(status="all", search="DID-P1-001", page=1, page_size=10)
            self.assertEqual(codes["items"][0]["status"], "available")

    def test_admin_request_ip_whitelist_api_reads_and_updates_database_config(self):
        with TemporaryDirectory() as temp_dir:
            settings = Settings(
                app_name="Test",
                host="127.0.0.1",
                port=18080,
                admin_username="admin",
                admin_password="password",
                secret_key="test-secret",
                data_dir=Path(temp_dir),
                request_ip_whitelist_enabled=True,
                request_ip_whitelist=("192.168.1.0/24",),
            )
            app = create_app(settings=settings)
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            initial_response = client.get("/api/admin/request-ip-whitelist")
            initial_payload = initial_response.get_json()

            self.assertEqual(initial_response.status_code, 200)
            self.assertTrue(initial_payload["success"])
            self.assertTrue(initial_payload["data"]["enabled"])
            self.assertEqual(initial_payload["data"]["allowed_ips"], ["192.168.1.0/24"])

            update_response = client.post(
                "/api/admin/request-ip-whitelist",
                json={
                    "enabled": True,
                    "allowed_ips": ["10.0.0.1", "10.0.0.0/24", "10.0.0.1"],
                },
            )
            update_payload = update_response.get_json()

            self.assertEqual(update_response.status_code, 200)
            self.assertTrue(update_payload["success"])
            self.assertEqual(
                update_payload["data"]["allowed_ips"],
                ["10.0.0.1", "10.0.0.0/24"],
            )

            reload_app = create_app(
                settings=Settings(
                    app_name="Test",
                    host="127.0.0.1",
                    port=18080,
                    admin_username="admin",
                    admin_password="password",
                    secret_key="test-secret",
                    data_dir=Path(temp_dir),
                    request_ip_whitelist_enabled=False,
                    request_ip_whitelist=(),
                )
            )
            configuration_service = reload_app.extensions["configuration_service"]
            self.assertEqual(
                configuration_service.get_request_ip_whitelist_config(),
                {
                    "enabled": True,
                    "allowed_ips": ["10.0.0.1", "10.0.0.0/24"],
                },
            )

    def test_excel_import_reports_all_invalid_rows_and_does_not_write_anything(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            excel_service = app.extensions["excel_service"]
            repo = app.extensions["auth_code_repository"]

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["did", "license", "source_batch"])
            sheet.append(["DID-001", "LIC-001", "B1"])
            sheet.append(["", "LIC-002", "B1"])
            sheet.append(["DID-003", "", "B1"])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            with self.assertRaises(Exception) as ctx:
                excel_service.import_codes(buffer, default_pid="PX")
            self.assertIn("第 3 行", ctx.exception.errors[0])
            self.assertIn("第 4 行", ctx.exception.errors[1])

            codes = repo.list_codes(status="all", search="PX", page=1, page_size=10)
            self.assertEqual(codes["total"], 0)

    def test_admin_import_api_returns_structured_errors(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["pid", "did", "license"])
            sheet.append(["PX", "", "LIC-001"])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            response = client.post(
                "/api/admin/import-codes",
                data={"file": (buffer, "invalid.xlsx")},
                content_type="multipart/form-data",
            )
            payload = response.get_json()

            self.assertEqual(response.status_code, 400)
            self.assertFalse(payload["success"])
            self.assertIn("导入校验失败", payload["message"])
            self.assertTrue(payload["errors"])
            self.assertIn("第 2 行", payload["errors"][0])

    def test_admin_logs_api_returns_recent_logs(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-LOG-001",
                        "license": "LIC-LOG-001",
                        "code": "DID-LOG-001",
                        "payload_json": "{\"did\": \"DID-LOG-001\", \"license\": \"LIC-LOG-001\"}",
                        "payload_hash": "log-hash-001",
                        "source_batch": "B1",
                    },
                ]
            )

            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            response = client.get("/api/admin/logs?limit=20")
            payload = response.get_json()

            self.assertEqual(response.status_code, 200)
            self.assertTrue(payload["success"])
            self.assertGreaterEqual(len(payload["data"]["items"]), 1)
            self.assertEqual(payload["data"]["items"][0]["pid"], "P1")
            self.assertIn("log_id", payload["data"]["items"][0])

    def test_admin_logs_api_supports_action_filter(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-LOG-001",
                        "license": "LIC-LOG-001",
                        "code": "DID-LOG-001",
                        "payload_json": "{\"did\": \"DID-LOG-001\", \"license\": \"LIC-LOG-001\"}",
                        "payload_hash": "log-hash-001",
                        "source_batch": "B1",
                    },
                ]
            )

            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-44-55-66", "pid": "PX"},
            )

            response = client.get("/api/admin/logs?limit=20&action=reused")
            payload = response.get_json()

            self.assertEqual(response.status_code, 200)
            self.assertTrue(payload["success"])
            self.assertEqual(payload["data"]["action"], "reused")
            self.assertEqual(len(payload["data"]["items"]), 1)
            self.assertEqual(payload["data"]["items"][0]["action"], "reused")

    def test_admin_logs_api_supports_pid_or_mac_keyword_filter(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-LOG-001",
                        "license": "LIC-LOG-001",
                        "code": "DID-LOG-001",
                        "payload_json": "{\"did\": \"DID-LOG-001\", \"license\": \"LIC-LOG-001\"}",
                        "payload_hash": "log-hash-001",
                        "source_batch": "B1",
                    },
                ]
            )

            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-44-55-66", "pid": "PX"},
            )

            response = client.get("/api/admin/logs?limit=20&search=44-55-66")
            payload = response.get_json()

            self.assertEqual(response.status_code, 200)
            self.assertTrue(payload["success"])
            self.assertEqual(payload["data"]["search"], "44-55-66")
            self.assertEqual(len(payload["data"]["items"]), 1)
            self.assertEqual(payload["data"]["items"][0]["pid"], "PX")
            self.assertEqual(payload["data"]["items"][0]["mac"], "AA:BB:CC:44:55:66")

    def test_admin_logs_export_uses_action_filter(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-EXPORT-001",
                        "license": "LIC-EXPORT-001",
                        "code": "DID-EXPORT-001",
                        "payload_json": "{\"did\": \"DID-EXPORT-001\", \"license\": \"LIC-EXPORT-001\"}",
                        "payload_hash": "export-hash-001",
                        "source_batch": "B1",
                    },
                ]
            )

            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-44-55-66", "pid": "PX"},
            )

            response = client.get("/api/admin/export-logs?action=reused")

            self.assertEqual(response.status_code, 200)
            self.assertIn(
                "request-logs-reused.xlsx",
                response.headers["Content-Disposition"],
            )

            workbook = load_workbook(BytesIO(response.data))
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            self.assertEqual(rows[0][:8], ("时间", "PID", "MAC", "动作", "DID", "说明", "来源IP", "载荷JSON"))
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][1], "P1")
            self.assertEqual(rows[1][3], "reused")
            self.assertEqual(rows[1][4], "DID-EXPORT-001")

    def test_admin_logs_export_supports_keyword_filter(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-EXPORT-SEARCH-001",
                        "license": "LIC-EXPORT-SEARCH-001",
                        "code": "DID-EXPORT-SEARCH-001",
                        "payload_json": "{\"did\": \"DID-EXPORT-SEARCH-001\", \"license\": \"LIC-EXPORT-SEARCH-001\"}",
                        "payload_hash": "export-search-hash-001",
                        "source_batch": "B1",
                    },
                ]
            )

            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-11-22-33", "pid": "P1"},
            )
            client.post(
                "/api/device/authorize",
                json={"mac": "AA-BB-CC-44-55-66", "pid": "PX"},
            )

            response = client.get("/api/admin/export-logs?search=44-55-66")

            self.assertEqual(response.status_code, 200)
            self.assertIn(
                "request-logs-all.xlsx",
                response.headers["Content-Disposition"],
            )

            workbook = load_workbook(BytesIO(response.data))
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][1], "PX")
            self.assertEqual(rows[1][2], "AA:BB:CC:44:55:66")
            self.assertEqual(rows[1][3], "exhausted")

    def test_admin_logs_export_supports_time_range_filter(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            database = app.extensions["database"]
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            with database.connection() as conn:
                conn.execute(
                    """
                    INSERT INTO distribution_logs (
                        pid, mac, code, payload_json, action, message, client_ip, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "P1",
                        "AA:BB:CC:11:22:33",
                        "DID-TIME-001",
                        "{\"did\": \"DID-TIME-001\", \"license\": \"LIC-TIME-001\"}",
                        "assigned",
                        "time-in-range",
                        "127.0.0.1",
                        "2026-05-25 10:15:00",
                    ),
                )
                conn.execute(
                    """
                    INSERT INTO distribution_logs (
                        pid, mac, code, payload_json, action, message, client_ip, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "P2",
                        "AA:BB:CC:44:55:66",
                        "DID-TIME-002",
                        "{\"did\": \"DID-TIME-002\", \"license\": \"LIC-TIME-002\"}",
                        "reused",
                        "time-out-of-range",
                        "127.0.0.1",
                        "2026-05-25 12:30:00",
                    ),
                )

            response = client.get(
                "/api/admin/export-logs?start_at=2026-05-25T10:00&end_at=2026-05-25T11:00"
            )

            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.data))
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][0], "2026-05-25 10:15:00")
            self.assertEqual(rows[1][1], "P1")
            self.assertEqual(rows[1][4], "DID-TIME-001")

    def test_admin_logs_export_rejects_invalid_time_range(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            client = app.test_client()
            client.post("/login", data={"username": "admin", "password": "password"})

            response = client.get(
                "/api/admin/export-logs?start_at=2026-05-25T12:00&end_at=2026-05-25T11:00"
            )
            payload = response.get_json()

            self.assertEqual(response.status_code, 400)
            self.assertFalse(payload["success"])
            self.assertIn("开始时间不能晚于结束时间", payload["message"])

    def test_excel_import_supports_same_did_under_different_pid(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            excel_service = app.extensions["excel_service"]
            repo = app.extensions["auth_code_repository"]

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["pid", "did", "license", "source_batch"])
            sheet.append(["P1", "DID-001", "LIC-001", "B1"])
            sheet.append(["P2", "DID-001", "LIC-002", "B2"])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            result = excel_service.import_codes(buffer)
            self.assertEqual(result["inserted"], 2)
            self.assertEqual(result["skipped"], 0)

            codes = repo.list_codes(status="all", search="DID-001", page=1, page_size=10)
            self.assertEqual(codes["total"], 2)

    def test_excel_import_warns_on_duplicate_pid_and_did(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            excel_service = app.extensions["excel_service"]
            repo = app.extensions["auth_code_repository"]

            initial_workbook = Workbook()
            initial_sheet = initial_workbook.active
            initial_sheet.append(["pid", "did", "license", "source_batch"])
            initial_sheet.append(["PX", "DID-001", "LIC-001", "B1"])
            initial_buffer = BytesIO()
            initial_workbook.save(initial_buffer)
            initial_buffer.seek(0)

            initial_result = excel_service.import_codes(initial_buffer)
            self.assertEqual(initial_result["inserted"], 1)

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["pid", "did", "license", "source_batch"])
            sheet.append(["PX", "DID-001", "LIC-001-NEW", "B1"])
            sheet.append(["PX", "DID-002", "LIC-002", "B1"])
            sheet.append(["PX", "DID-002", "LIC-002-DUP", "B1"])
            sheet.append(["PY", "DID-001", "LIC-003", "B2"])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            result = excel_service.import_codes(buffer)
            self.assertEqual(result["inserted"], 2)
            self.assertEqual(result["skipped"], 2)
            self.assertIn("PX / DID-001", result["duplicate_records"])
            self.assertIn("PX / DID-002", result["duplicate_records"])
            self.assertTrue(result["warnings"])

            codes = repo.list_codes(status="all", search="DID-001", page=1, page_size=10)
            self.assertEqual(codes["total"], 2)

    def test_inventory_summary_is_grouped_by_pid(self):
        with TemporaryDirectory() as temp_dir:
            app = self.create_test_app(temp_dir)
            repo = app.extensions["auth_code_repository"]
            repo.bulk_insert_codes(
                [
                    {
                        "pid": "P1",
                        "did": "DID-SUMMARY-P1-1",
                        "license": "P1-001",
                        "code": "DID-SUMMARY-P1-1",
                        "payload_json": "{\"did\": \"DID-SUMMARY-P1-1\", \"license\": \"P1-001\"}",
                        "payload_hash": "summary-hash-p1-1",
                        "source_batch": "B1",
                    },
                    {
                        "pid": "P1",
                        "did": "DID-SUMMARY-P1-2",
                        "license": "P1-002",
                        "code": "DID-SUMMARY-P1-2",
                        "payload_json": "{\"did\": \"DID-SUMMARY-P1-2\", \"license\": \"P1-002\"}",
                        "payload_hash": "summary-hash-p1-2",
                        "source_batch": "B1",
                    },
                    {
                        "pid": "P2",
                        "did": "DID-SUMMARY-P2-1",
                        "license": "P2-001",
                        "code": "DID-SUMMARY-P2-1",
                        "payload_json": "{\"did\": \"DID-SUMMARY-P2-1\", \"license\": \"P2-001\"}",
                        "payload_hash": "summary-hash-p2-1",
                        "source_batch": "B2",
                    },
                ]
            )

            summary = repo.summarize_inventory_by_pid(page=1, page_size=10)
            self.assertEqual(summary["total"], 2)
            self.assertEqual(summary["items"][0]["pid"], "P1")
            self.assertEqual(summary["items"][0]["total_codes"], 2)
